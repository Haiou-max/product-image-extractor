#!/usr/bin/env python3
"""
P-E 飞书机器人 — 发送产品图片，自动提取货号/内装/单价/尺寸，返回 Excel。
使用 WebSocket 长连接模式，无需公网回调地址。
"""

import anthropic
import base64
import json
import re
import os
import io
import tempfile
import logging
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import lark_oapi as lark
from lark_oapi.api.im.v1 import *

# --- 配置 ---
FEISHU_APP_ID = "cli_a95de4b803fa9cc8"
FEISHU_APP_SECRET = "6b7ehCMDATHUgox2EsMkmfjuoqcEUKbW"

EVOLINK_API_KEY = "sk-XtjGuMeJV15fcrQRKXirDvSzg3M9wBoYQCmRiSQ0v9WEt52A"
EVOLINK_BASE_URL = "https://api.evolink.ai"
EVOLINK_MODEL = "claude-sonnet-4-20250514"

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# --- 飞书客户端 ---
lark_client = lark.Client.builder().app_id(FEISHU_APP_ID).app_secret(FEISHU_APP_SECRET).build()

# --- Claude 客户端 ---
claude_client = anthropic.Anthropic(api_key=EVOLINK_API_KEY, base_url=EVOLINK_BASE_URL)

SYSTEM_PROMPT = """你是一个产品信息提取助手。用户会发送一张产品照片，照片中有产品实物和手写的产品信息。

请从图片中提取以下 4 个字段，严格返回 JSON 格式（不要返回其他内容）：

1. **货号** - 格式为 X-X（如 3-6, 3-15），通常是第一行手写内容
2. **内装** - 格式为 X pcs 或 Xpcs（如 12 pcs, 36pcs），只提取数字
3. **单价** - 格式为 X ¥ 或 X￥（如 55 ¥, 18.8¥），只提取数字
4. **尺寸** - 格式为 X × X 或 XxX（如 30.5×24, 14x19.5），保留原样

返回格式：
{"货号": "3-15", "内装": 12, "单价": 33, "尺寸": "24×12"}

注意：
- 手写字体可能不规范，请结合上下文推断
- 如果某个字段无法识别，值设为 null
- 单价可能是小数（如 18.8）
- 只返回 JSON，不要其他文字"""


# --- 图片识别 ---
def extract_from_image(image_bytes, media_type="image/jpeg"):
    """调用 Claude API 从图片中提取产品信息。"""
    b64 = base64.standard_b64encode(image_bytes).decode("utf-8")
    response = claude_client.messages.create(
        model=EVOLINK_MODEL,
        max_tokens=256,
        system=SYSTEM_PROMPT,
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
                    {"type": "text", "text": "请提取这张图片中的产品信息。"},
                ],
            }
        ],
    )
    text = response.content[0].text.strip()
    json_match = re.search(r"\{.*\}", text, re.DOTALL)
    if json_match:
        return json.loads(json_match.group())
    return json.loads(text)


# --- Excel 生成 ---
def generate_excel(results, image_bytes_list):
    """生成带嵌入图片的 Excel 文件，返回文件路径。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "产品列表"

    headers = ["图片", "货号", "内装", "单价", "尺寸"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for i, (data, img_bytes) in enumerate(zip(results, image_bytes_list), 2):
        ws.cell(row=i, column=2, value=data.get("货号", "")).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=i, column=3, value=data.get("内装", "")).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=i, column=4, value=data.get("单价", "")).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=i, column=5, value=data.get("尺寸", "")).alignment = Alignment(horizontal="center", vertical="center")
        for col in range(1, 6):
            ws.cell(row=i, column=col).border = thin_border

        img = XlImage(io.BytesIO(img_bytes))
        img.width = 320
        img.height = 240
        ws.add_image(img, f"A{i}")
        ws.row_dimensions[i].height = 190

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 18
    ws.row_dimensions[1].height = 25

    output_path = os.path.join(tempfile.gettempdir(), "product_list.xlsx")
    wb.save(output_path)
    return output_path


# --- 飞书消息操作 ---
def get_tenant_access_token():
    """获取 tenant_access_token。"""
    resp = requests.post(
        "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
        json={"app_id": FEISHU_APP_ID, "app_secret": FEISHU_APP_SECRET},
    )
    return resp.json().get("tenant_access_token")


def download_image(message_id, image_key):
    """从飞书下载图片。"""
    token = get_tenant_access_token()
    resp = requests.get(
        f"https://open.feishu.cn/open-apis/im/v1/messages/{message_id}/resources/{image_key}",
        params={"type": "image"},
        headers={"Authorization": f"Bearer {token}"},
    )
    if resp.status_code == 200:
        return resp.content
    logger.error(f"下载图片失败: {resp.status_code} {resp.text}")
    return None


def upload_file(file_path):
    """上传文件到飞书，返回 file_key。"""
    token = get_tenant_access_token()
    with open(file_path, "rb") as f:
        resp = requests.post(
            "https://open.feishu.cn/open-apis/im/v1/files",
            headers={"Authorization": f"Bearer {token}"},
            data={"file_type": "stream", "file_name": "product_list.xlsx"},
            files={"file": ("product_list.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    data = resp.json()
    if data.get("code") == 0:
        return data["data"]["file_key"]
    logger.error(f"上传文件失败: {data}")
    return None


def send_text(chat_id, text):
    """发送文本消息。"""
    request = CreateMessageRequest.builder() \
        .receive_id_type("chat_id") \
        .request_body(
            CreateMessageRequestBody.builder()
            .receive_id(chat_id)
            .msg_type("text")
            .content(json.dumps({"text": text}))
            .build()
        ).build()
    lark_client.im.v1.message.create(request)


def send_file(chat_id, file_key):
    """发送文件消息。"""
    request = CreateMessageRequest.builder() \
        .receive_id_type("chat_id") \
        .request_body(
            CreateMessageRequestBody.builder()
            .receive_id(chat_id)
            .msg_type("file")
            .content(json.dumps({"file_key": file_key}))
            .build()
        ).build()
    lark_client.im.v1.message.create(request)


# --- 消息缓存（支持多图批量处理） ---
# chat_id -> {"images": [(message_id, image_key), ...], "timer": None}
pending_images = {}


def process_and_reply(chat_id, message_id, image_key):
    """处理单张图片并回复。"""
    logger.info(f"开始处理图片: chat={chat_id}, image_key={image_key}")
    send_text(chat_id, "正在识别产品信息，请稍候...")

    image_bytes = download_image(message_id, image_key)
    if not image_bytes:
        send_text(chat_id, "下载图片失败，请重新发送。")
        return

    try:
        data = extract_from_image(image_bytes)
        logger.info(f"识别结果: {data}")

        # 生成文字结果
        lines = [
            f"货号: {data.get('货号', '未识别')}",
            f"内装: {data.get('内装', '未识别')}",
            f"单价: {data.get('单价', '未识别')}",
            f"尺寸: {data.get('尺寸', '未识别')}",
        ]
        send_text(chat_id, "识别完成\n" + "\n".join(lines))

        # 生成 Excel 并发送
        excel_path = generate_excel([data], [image_bytes])
        file_key = upload_file(excel_path)
        if file_key:
            send_file(chat_id, file_key)
        else:
            send_text(chat_id, "Excel 文件上传失败，请重试。")

    except Exception as e:
        logger.error(f"处理失败: {e}")
        send_text(chat_id, f"识别失败: {e}")


# --- 飞书事件处理 ---
def handle_message(data: lark.RawEvent):
    """处理收到的消息事件。"""
    try:
        event = json.loads(data.raw)
        header = event.get("header", {})
        event_data = event.get("event", {})
        message = event_data.get("message", {})

        event_type = header.get("event_type", "")
        if event_type != "im.message.receive_v1":
            return

        msg_type = message.get("message_type", "")
        chat_id = message.get("chat_id", "")
        message_id = message.get("message_id", "")

        if msg_type == "image":
            content = json.loads(message.get("content", "{}"))
            image_key = content.get("image_key", "")
            if image_key and chat_id:
                process_and_reply(chat_id, message_id, image_key)

        elif msg_type == "text":
            content = json.loads(message.get("content", "{}"))
            text = content.get("text", "").strip()
            if text in ("帮助", "help", "/help"):
                send_text(
                    chat_id,
                    "发送产品图片，我会自动识别货号/内装/单价/尺寸并生成 Excel 表格。\n\n"
                    "直接发图片即可，无需其他操作。",
                )
            else:
                send_text(chat_id, "请发送产品图片，我会自动识别产品信息并生成 Excel。")

    except Exception as e:
        logger.error(f"事件处理异常: {e}", exc_info=True)


def main():
    """启动飞书机器人。"""
    logger.info("P-E 飞书机器人启动中...")

    event_handler = lark.EventDispatcherHandler.builder("", "") \
        .register_raw_event(lark.RawEvent, handle_message) \
        .build()

    cli = lark.ws.Client(
        FEISHU_APP_ID,
        FEISHU_APP_SECRET,
        event_handler=event_handler,
        log_level=lark.LogLevel.INFO,
    )

    logger.info("机器人已启动，等待消息...")
    cli.start()


if __name__ == "__main__":
    main()
