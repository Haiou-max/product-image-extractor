import streamlit as st
import anthropic
import base64
import json
import io
import re
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

st.set_page_config(page_title="产品图片提取", page_icon="📦", layout="wide")

st.title("📦 产品图片 → Excel 提取工具")
st.markdown("上传产品照片，自动识别 **货号 / 内装 / 单价 / 尺寸**，一键生成 Excel 表格。")

SYSTEM_PROMPT = """你是一个产品信息提取助手。用户会发送一张产品照片，照片中有产品实物和手写的产品信息。

请从图片中提取以下 4 个字段，严格返回 JSON 格式（不要返回其他内容）：

1. **货号** - 格式为 X-X（如 3-6, 3-15），通常是第一行手写内容
2. **内装** - 格式为 X pcs 或 Xpcs（如 12 pcs, 36pcs），只提取数字
3. **单价** - 格式为 X ¥ 或 X￥（如 55 ¥, 18.8¥），只提取数字
4. **尺寸** - 格式为 X × X 或 XxX（如 30.5×24, 14x19.5），保留原样

返回格式：
{"货号": "3-15", "内装": 12, "单价": 33, "尺寸": "24×12"}

注意：
- 手写字体可能不规范，请结合上下文推断
- 如果某个字段无法识别，值设为 null
- 单价可能是小数（如 18.8）
- 只返回 JSON，不要其他文字"""


def extract_from_image(client, image_bytes, media_type):
    b64 = base64.standard_b64encode(image_bytes).decode("utf-8")
    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=256,
        system=SYSTEM_PROMPT,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": b64,
                        },
                    },
                    {"type": "text", "text": "请提取这张图片中的产品信息。"},
                ],
            }
        ],
    )
    text = response.content[0].text.strip()
    # Extract JSON from possible markdown code block
    json_match = re.search(r"\{.*\}", text, re.DOTALL)
    if json_match:
        return json.loads(json_match.group())
    return json.loads(text)


def generate_excel(results, uploaded_files):
    wb = Workbook()
    ws = wb.active
    ws.title = "产品列表"

    headers = ["图片", "货号", "内装", "单价", "尺寸"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for i, (data, file) in enumerate(zip(results, uploaded_files), 2):
        ws.cell(row=i, column=2, value=data.get("货号", "")).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(row=i, column=3, value=data.get("内装", "")).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(row=i, column=4, value=data.get("单价", "")).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(row=i, column=5, value=data.get("尺寸", "")).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        for col in range(1, 6):
            ws.cell(row=i, column=col).border = thin_border

        # Embed image
        img_bytes = file.getvalue()
        img = XlImage(io.BytesIO(img_bytes))
        img.width = 320
        img.height = 240
        ws.add_image(img, f"A{i}")
        ws.row_dimensions[i].height = 190

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 18
    ws.row_dimensions[1].height = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# --- UI ---

uploaded_files = st.file_uploader(
    "上传产品图片",
    type=["jpg", "jpeg", "png"],
    accept_multiple_files=True,
    help="支持 JPG / PNG 格式，可同时上传多张",
)

if uploaded_files:
    st.markdown(f"已上传 **{len(uploaded_files)}** 张图片：")
    cols = st.columns(min(len(uploaded_files), 4))
    for idx, file in enumerate(uploaded_files):
        with cols[idx % 4]:
            st.image(file, caption=file.name, width=200)

    if st.button("🚀 开始提取", type="primary", use_container_width=True):
        api_key = st.secrets.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            st.error("未配置 ANTHROPIC_API_KEY，请在 Streamlit Secrets 中添加。")
            st.stop()

        client = anthropic.Anthropic(api_key=api_key)
        results = []
        progress = st.progress(0, text="正在识别...")

        for idx, file in enumerate(uploaded_files):
            progress.progress(
                (idx + 1) / len(uploaded_files),
                text=f"正在识别 {file.name} ({idx + 1}/{len(uploaded_files)})",
            )
            file.seek(0)
            img_bytes = file.read()
            media_type = "image/jpeg" if file.name.lower().endswith((".jpg", ".jpeg")) else "image/png"
            try:
                data = extract_from_image(client, img_bytes, media_type)
                results.append(data)
            except Exception as e:
                st.warning(f"⚠️ {file.name} 识别失败: {e}")
                results.append({"货号": "[识别失败]", "内装": None, "单价": None, "尺寸": None})

        progress.empty()

        if results:
            st.success(f"识别完成！共 {len(results)} 条数据")

            # Show results table
            st.dataframe(
                [
                    {
                        "货号": r.get("货号", ""),
                        "内装": r.get("内装", ""),
                        "单价": r.get("单价", ""),
                        "尺寸": r.get("尺寸", ""),
                    }
                    for r in results
                ],
                use_container_width=True,
            )

            # Generate and download Excel
            for file in uploaded_files:
                file.seek(0)
            excel_bytes = generate_excel(results, uploaded_files)
            st.download_button(
                label="📥 下载 Excel 文件",
                data=excel_bytes,
                file_name="product_list.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
